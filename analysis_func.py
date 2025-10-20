import numpy as np
# import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
import matplotlib.pyplot as plt

def split_into_q2_bins(data, bin_bounds=[0, 1, 2, 3, 4, 5, 6, 7, 8, 11, 12.5, 15, 17, 19, 23]):
    data['dimuon_mass_squared'] = (data['dimuon_system_invariant_mass'] / 1000) ** 2
    data["q2_bin"] = pd.cut(data["dimuon_mass_squared"], bins=bin_bounds, right=False, labels=False)

    # Separate DataFrames per bin
    binned_data = {}

    for i in range(len(bin_bounds) - 1):
        if bin_bounds[i] == 8:
            binned_resonant_data = data[data["q2_bin"] == i].copy()
        elif bin_bounds[i] == 12.5:
            continue
        else:
            binned_data[i] = data[data["q2_bin"] == i].copy()

    binned_resonant_data = binned_resonant_data[(abs(binned_resonant_data['B_invariant_mass'] - 5280) < 150) & (abs(binned_resonant_data['dimuon_system_invariant_mass'] - 3097) < 50)]
    return binned_data, binned_resonant_data, bin_bounds

def post_selection_vetoes(fs_data, diagnostics=False, visual=False):
    # Detector Resolution for Invariant Mass = 8MeV (stddev)
    m_jpsik, m_psi2s, veto_window = 3096.916, 3686.109, 60 # MeV
    data, datap = [], []
    fs_data = fs_data[(fs_data['B_invariant_mass'] < 5700) & (fs_data['B_invariant_mass'] > 5000)].copy() # reduce candidate mass range
    fs_data = fs_data[fs_data['Kaon_PID_NN_score_for_kaon_hypothesis'] > 0.4] # removes JPsiK decays where the muon was misidentified as kaon


    # D meson - Pion misidentified as Muon (1864.84+-0.17 MeV/c2)
    fs_data['kmu_mass'] = kmu_calc(fs_data['Kaon_4_momentum_x_component'], fs_data['Kaon_4_momentum_y_component'],
                                   fs_data['Kaon_4_momentum_z_component'], fs_data['Opposite_sign_muon_4_momentum_x_component'],
                                   fs_data['Opposite_sign_muon_4_momentum_y_component'], fs_data['Opposite_sign_muon_4_momentum_z_component'],
                                   mu_pi_hypo=True)
    data.append(fs_data['kmu_mass'].copy())
    fs_data = fs_data[(fs_data['kmu_mass'] < 1845) | (fs_data['kmu_mass'] > 1885)] # 98.76% removal |
    datap.append(fs_data['kmu_mass'].copy())


    # Hadron misidentification as Muon, only removes 13 out of 3000+ events, debatable whether to keep or not
    fs_data['kmu_mass'] = kmu_calc(fs_data['Kaon_4_momentum_x_component'], fs_data['Kaon_4_momentum_y_component'],
                                    fs_data['Kaon_4_momentum_z_component'], fs_data['Opposite_sign_muon_4_momentum_x_component'],
                                    fs_data['Opposite_sign_muon_4_momentum_y_component'], fs_data['Opposite_sign_muon_4_momentum_z_component'],
                                    k_mu_hypo=True)
    data.append(fs_data['kmu_mass'].copy())
    mask_keep = ~((fs_data['Kaon_PID_NN_score_for_muon_hypothesis'] > -100) &
                  ((np.abs(fs_data['kmu_mass'] - m_jpsik) < veto_window) | (np.abs(fs_data['kmu_mass'] - m_psi2s) < veto_window)))
    fs_data = fs_data[mask_keep]
    # fs_data = fs_data[(abs(fs_data['kmu_mass'] - 1680) > 40) & (fs_data['Opposite_sign_muon_PID_NN_score_for_kaon_hypothesis'] < 0.5)]
    datap.append(fs_data['kmu_mass'].copy())


    # Resonant decays where Kaon and Muon misidentified as each other - https://journals.aps.org/prl/pdf/10.1103/PhysRevLett.111.151801
    fs_data['kmu_mass'] = kmu_calc(fs_data['Kaon_4_momentum_x_component'], fs_data['Kaon_4_momentum_y_component'],
                                   fs_data['Kaon_4_momentum_z_component'], fs_data['Opposite_sign_muon_4_momentum_x_component'],
                                   fs_data['Opposite_sign_muon_4_momentum_y_component'], fs_data['Opposite_sign_muon_4_momentum_z_component'],
                                   kmu_swap_hypo=True)
    data.append(fs_data['kmu_mass'].copy())
    fs_data = fs_data[(fs_data['kmu_mass'] < 3075) | (fs_data['kmu_mass'] > 3118)] # 99.27% removal |
    datap.append(fs_data['kmu_mass'].copy())


    # Diagnostics
    if diagnostics:
        res = []
        for d, dp in zip(data, datap):
            res.append({
            "yield_pre_veto": len(d),
            "yield_post_veto": len(dp),
            "fraction_kept": (len(d) - len(dp)) / len(d),
            })
            if visual:
                plt.hist(d, bins=50)
                plt.hist(dp, bins=50)
                plt.grid()
                plt.ylabel('Yields')
                plt.xlabel('Invariant Mass (MeV)')
                plt.show()
        print('##### DISPLAYING DIAGNOSTICS FOR POST-SELECTION VETOES #####')
        print(pd.DataFrame(res).to_string(index=False))
    return fs_data

def kmu_calc(k_px, k_py, k_pz, mu_px, mu_py, mu_pz, mu_pi_hypo=False, k_mu_hypo=False, kmu_swap_hypo=False):
    k_pm_mass = 493.677 # +-0.016 MeV/c2
    mu_mass = 105.658 # MeV/c2
    pi_pm_mass = 139.570 # +-0.00018 MeV/c2

    p1_mass = k_pm_mass
    p2_mass = mu_mass
    if mu_pi_hypo:
        p2_mass = pi_pm_mass
    elif k_mu_hypo:
        p1_mass = mu_mass
    elif kmu_swap_hypo:
        p1_mass = mu_mass
        p2_mass = k_pm_mass
    k_E = np.sqrt(k_px**2 + k_py**2 + k_pz**2 + p1_mass**2)
    mu_E = np.sqrt(mu_px**2 + mu_py**2 + mu_pz**2 + p2_mass**2)
    E_tot = k_E + mu_E
    px_tot = k_px + mu_px
    py_tot = k_py + mu_py
    pz_tot = k_pz + mu_pz
    inv_mass = E_tot**2 - (px_tot**2 + py_tot**2 + pz_tot**2)
    return np.sqrt(np.maximum(inv_mass, 0.0))

def kmu_mass_muon_hypothesis(k_px, k_py, k_pz, mu_px, mu_py, mu_pz):
    m_mu = 105.6583755  # MeV
    # recompute kaon energy under muon mass
    k_E = np.sqrt(k_px**2 + k_py**2 + k_pz**2 + m_mu**2)
    mu_E = np.sqrt(mu_px**2 + mu_py**2 + mu_pz**2 + m_mu**2)
    E_tot = k_E + mu_E
    px_tot = k_px + mu_px
    py_tot = k_py + mu_py
    pz_tot = k_pz + mu_pz
    inv_mass2 = E_tot**2 - (px_tot**2 + py_tot**2 + pz_tot**2)
    return np.sqrt(np.maximum(inv_mass2, 0.0))

def jpsik_swapped_hypothesis(k_px, k_py, k_pz, mu_px, mu_py, mu_pz):
    m_mu = 105.6583755  # MeV
    m_k = 493.677 # MeV
    # recompute kaon energy under muon mass
    k_E = np.sqrt(k_px**2 + k_py**2 + k_pz**2 + m_mu**2)
    mu_E = np.sqrt(mu_px**2 + mu_py**2 + mu_pz**2 + m_k**2)
    E_tot = k_E + mu_E
    px_tot = k_px + mu_px
    py_tot = k_py + mu_py
    pz_tot = k_pz + mu_pz
    inv_mass2 = E_tot**2 - (px_tot**2 + py_tot**2 + pz_tot**2)
    return np.sqrt(np.maximum(inv_mass2, 0.0))

def kmu_mass_filter(fs_data):
    # fs_data = fs_data[(fs_data['B_invariant_mass'] < 5500) & (fs_data['B_invariant_mass'] > 5000)].copy()
    fs_data['kmu_mass'] = kmu_calc(fs_data['Kaon_4_momentum_energy_component'], fs_data['Opposite_sign_muon_4_momentum_energy_component'],
                                fs_data['Kaon_4_momentum_x_component'], fs_data['Kaon_4_momentum_y_component'],
                                fs_data['Kaon_4_momentum_z_component'], fs_data['Opposite_sign_muon_4_momentum_x_component'],
                                fs_data['Opposite_sign_muon_4_momentum_y_component'], fs_data['Opposite_sign_muon_4_momentum_z_component'])
    fs_data = fs_data[(fs_data['kmu_mass'] < 1845) | (fs_data['kmu_mass'] > 1885)] # misidentified D mesons (98.76% removed |)
    return fs_data

def veto_filter(fs_dataa):
    m_jpsik = 3096.916
    m_psi2s = 3686.109
    veto_window = 50

    # veto hadron misidentification as muon
    fs_data = fs_dataa.copy()
    fs_data['kmu_mass'] = kmu_mass_muon_hypothesis(fs_data['Kaon_4_momentum_x_component'], fs_data['Kaon_4_momentum_y_component'],
    fs_data['Kaon_4_momentum_z_component'], fs_data['Opposite_sign_muon_4_momentum_x_component'],
    fs_data['Opposite_sign_muon_4_momentum_y_component'], fs_data['Opposite_sign_muon_4_momentum_z_component'])
    mask_keep = ~(
    (fs_data['Kaon_PID_NN_score_for_muon_hypothesis'] > -100) &
    ((np.abs(fs_data['kmu_mass'] - m_jpsik) < veto_window) |
     (np.abs(fs_data['kmu_mass'] - m_psi2s) < veto_window)))
    fs_data = fs_data[mask_keep].copy() # only removes 13 out of 3000+ events, debatable whether to keep or not

    # remove resonance from hadron and muon being misidentified as each other (swap)
    fs_data['kmu_mass'] = jpsik_swapped_hypothesis(fs_data['Kaon_4_momentum_x_component'], fs_data['Kaon_4_momentum_y_component'],
    fs_data['Kaon_4_momentum_z_component'], fs_data['Opposite_sign_muon_4_momentum_x_component'],
    fs_data['Opposite_sign_muon_4_momentum_y_component'], fs_data['Opposite_sign_muon_4_momentum_z_component'])
    fs_data = fs_data[abs(fs_data['kmu_mass'] - m_jpsik) > veto_window]
    # Keep only tracks that are muon-like and NOT pion-like
    # mask_muon = (fs_data['Opposite_sign_muon_PID_NN_score_for_pion_hypothesis'] < 0.8) # pion PID threshold, large spread
    # fs_data = fs_data[mask_muon]
    return fs_data

def acp_calc(pd1, pd2):
    n1 = len(pd1)
    n2 = len(pd2)
    acp = (n1 - n2) / (n1 + n2)
    err = 2 / (n2 + n1)**2 * np.sqrt((n1 * np.sqrt(n2))**2 + (n2 * np.sqrt(n1))**2)
    return acp, err

def get_val_err(result, param):
        info = result.params[param]
        val = info["value"]
        err = (
            info.get("error")
            or (info.get("hesse") or {}).get("error")
            or 0.0
        )
        return val, err

def colour(dir, o_dir):
    df = pd.read_excel(dir, index_col=0)
    wb = load_workbook(dir)
    ws = wb.active   # assuming the first sheet

    n_rows = df.shape[0]
    n_cols = df.shape[1]

    start_cell = ws.cell(row=2, column=2).coordinate
    end_cell = ws.cell(row=1 + n_rows, column=1 + n_cols).coordinate
    cell_range = f"{start_cell}:{end_cell}"

    color_scale = ColorScaleRule(
        start_type='num', start_value=0, start_color='ffffff',   # green
        end_type='num',   end_value=1, end_color='FF0000'        # red
    )
    ws.conditional_formatting.add(cell_range, color_scale)
    # Save the workbook
    wb.save(o_dir)

# fs_data = pd.read_pickle("/Users/zifei/Desktop/C1/ProcessedData/first_selection_data2.pkl")
# plt.hist((kmu_mass_filter(fs_data)['kmu_mass']), bins=50)
# plt.grid()
# # plt.savefig('/Users/zifei/Desktop/kmu_minus.png', dpi=300)
# plt.show()